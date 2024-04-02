""" Small cli helper tool to replace patient names with patient ids in excel files.

This script is used to replace patient names with patient ids in excel files.
The script will look for excel files in the source directory and replace the names
with the ids. The script will create a new directory called 'output' next to the source directory 
and save the changed files there.
The source directory is specified by the user via prompt. 
The output directory is specified by the user via option --output, and defaults to "output".
The script is logging to a file called 'main_replace_patient_names.log' 
in a subdirectory called 'logs' inside the output directory.

Example call:
    $ python main.py --src /path/to/excel_files --output output
"""

import logging
import shutil
from pathlib import Path
from zipfile import BadZipFile

import click
import openpyxl
import pandas as pd

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("a4d-replacer-tool")

PATIENT_DATA_RANGE = ("B1", "C200")


@click.command("a4d-replacer-tool")
@click.option(
    "--src",
    "-s",
    prompt="Source directory",
    default=lambda: Path("."),
    help="Source directory",
)
@click.option("--output", "-o", default="output", help="Output directory")
@click.version_option("0.1.0", prog_name="replace patient names")
def replace_name_with_id(src: Path, output: str):
    """Replaces patient name with patient id in all excel files found in src.

    Args:
        src (Path): Source directory holding excel files as xlsx.
        output (str): Output directory holding excel files with replaced names.

    Raises:
        ValueError: If no excel files are found in src.
    """
    excel_files = list(Path(src).rglob("*.xlsx"))

    if not excel_files:
        raise ValueError(f"No excel files found in directory {src}.")

    output_dir = Path(src).parent / output / "patient_data_without_names"
    log_dir = Path(src).parent / output / "logs"

    if not output_dir.exists():
        output_dir.mkdir(parents=True)
        logger.info("Created output directory %s under %s.", output_dir, src)

    if not log_dir.exists():
        log_dir.mkdir(parents=True)
        logger.info("Created log directory %s under %s.", log_dir, src)

    handler = logging.FileHandler(log_dir / "main_replace_patient_names.log")
    handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(handler)

    logger.info("Start processing %s excel files.", len(excel_files))
    for i, excel_file in enumerate(excel_files):
        logger.info(
            "Start processing %s (%s/%s).", excel_file.name, i + 1, len(excel_files)
        )

        try:
            wb = openpyxl.load_workbook(str(excel_file), data_only=True)
        except BadZipFile:
            logger.warning("Failed to open %s. Skipping.", excel_file.name)
            continue

        sheets = wb.sheetnames

        if "Patient List" not in sheets:
            logger.warning("Sheet 'Patient List' not found. Skipping.")
            # shutil.copy(excel_file, output_dir / excel_file.name)
            continue

        patient_data = pd.DataFrame(
            [
                (b.value, c.value)
                for b, c in wb["Patient List"][
                    PATIENT_DATA_RANGE[0] : PATIENT_DATA_RANGE[1]
                ]
                if b.value and c.value
            ],
            columns=["id", "name"],
        )
        patient_data = patient_data.iloc[1:]
        patients_replaced = {name: False for name in patient_data.name}

        if all(patient_data.name == patient_data.id):
            logger.info(
                "Patient names are already replaced with ids. Copying original file to output dir."
            )
            shutil.copy(excel_file, output_dir / excel_file.name)
            continue

        for patient in patient_data.itertuples(index=False):
            for sheet in sheets:
                ws = wb[sheet]

                for col in ws.iter_cols():
                    for cell in col:
                        if cell.value == patient.name:
                            if sheet != "Patient List":
                                patients_replaced[patient.name] = True
                            cell.value = patient.id

        if not all(patients_replaced.values()):
            logger.warning(
                "Not all patient names were replaced. Missing patients: %s",
                ", ".join(
                    name for name, replaced in patients_replaced.items() if not replaced
                ),
            )

        wb.save(output_dir / excel_file.name)
        logger.info("Saved changed file to %s.", output_dir)
        logger.info(f"Finished processing %s.", excel_file.name)


if __name__ == "__main__":
    replace_name_with_id()
